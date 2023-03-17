from zipfile import ZipFile, ZIP_DEFLATED
import os
import csv
import PyPDF2
import openpyxl


def check_row_num_in_xlsx():
    xlsxfile = openpyxl.load_workbook('examples/sample3.xlsx')
    sheet = xlsxfile.active
    xlsx_rows = sheet.max_row
    return xlsx_rows


def check_row_num_in_csv():
    with open("examples/cities.csv", 'r', newline='') as csvfile:
        csv_reader = csv.reader(csvfile, delimiter=',')
        csv_rows = 0
        for row in csv_reader:
            csv_rows += 1
        return csv_rows


def check_page_num_in_pdf():
    with open("examples/sample.pdf", "rb") as pdffile:
        pdf_reader = PyPDF2.PdfReader(pdffile)
        pdf_numpages = len(pdf_reader.pages)
        return pdf_numpages


def move_files_into_archive():
    with ZipFile("resourses/myzip.zip", "w", compression=ZIP_DEFLATED, compresslevel=5) as myzip:
        myzip.write("examples/cities.csv", "cities.csv")
        cities_size_in_archive = myzip.getinfo("cities.csv").file_size
        cities = myzip.open("cities.csv")
        csv_rows_in_archive = 0
        for row in cities:
            csv_rows_in_archive += 1
        myzip.write("examples/sample.pdf", "sample.pdf")
        pdffile = myzip.open("sample.pdf")
        pdf_reader = PyPDF2.PdfReader(pdffile)
        pdf_numpages_in_archive = len(pdf_reader.pages)
        sample_size_in_archive = myzip.getinfo("sample.pdf").file_size
        myzip.write("examples/sample3.xlsx", "sample3.xlsx")
        xlsxfile = myzip.open("sample3.xlsx")
        xlsx_reader = openpyxl.load_workbook(xlsxfile)
        sheet = xlsx_reader.active
        xsls_rows_in_archive = sheet.max_row
        sample3_size_in_archive = myzip.getinfo("sample3.xlsx").file_size
        return [cities_size_in_archive, pdf_numpages_in_archive, sample_size_in_archive, \
                xsls_rows_in_archive, sample3_size_in_archive, csv_rows_in_archive]


cities_size = os.path.getsize("examples/cities.csv")
sample_size = os.path.getsize("examples/sample.pdf")
sample3_size = os.path.getsize("examples/sample3.xlsx")


def test_check_files_in_archive():
    rez = move_files_into_archive()
    cities_size_in_archive = rez[0]
    pdf_numpages_in_archive = rez[1]
    sample_size_in_archive = rez[2]
    xsls_rows_in_archive = rez[3]
    sample3_size_in_archive = rez[4]
    csv_rows_in_archive = rez[5]

    assert cities_size == cities_size_in_archive, f"Размер файла в архиве не совпадает с тем, что был до архивирования"
    assert sample_size == sample_size_in_archive, f"Размер файла в архиве не совпадает с тем, что был до архивирования"
    assert sample3_size == sample3_size_in_archive, f"Размер файла в архиве не совпадает с тем, что был до архивирования"
    assert check_row_num_in_csv() == csv_rows_in_archive, f"Количество строк csv файла в архиве не совпадает с изначальным кол-вом"
    assert check_page_num_in_pdf() == pdf_numpages_in_archive, f"Количество страниц pdf файла в архиве не совпадает с изначальным кол-вом"
    assert check_row_num_in_xlsx() == xsls_rows_in_archive, f"Количество строк на активном листе файла xlsx в архиве не совпадает с изначальным кол-вом"
